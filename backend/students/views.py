from rest_framework import viewsets, permissions, filters
from rest_framework.decorators import action
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from .models import Student
from .serializers import StudentSerializer


class StudentViewSet(viewsets.ModelViewSet):
    serializer_class = StudentSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'admission_no', 'grade', 'division', 'parent_name']

    def get_queryset(self):
        qs = Student.objects.all()

        params = self.request.query_params
        grade = params.get('grade')
        division = params.get('division')
        gender = params.get('gender')

        if grade:
            qs = qs.filter(grade=grade)
        if division:
            qs = qs.filter(division=division)
        if gender:
            qs = qs.filter(gender=gender)

        return qs

    @action(detail=False, methods=['get'], url_path='export')
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        wb = Workbook()
        ws = wb.active
        ws.title = "Students"

        headers = [
            "Name", "Admission No", "Grade", "Division",
            "Date of Birth", "Gender", "Parent Name", "Contact Number",
        ]
        ws.append(headers)

        for s in queryset:
            ws.append([
                s.name,
                s.admission_no,
                s.grade,
                s.division,
                s.dob.isoformat() if s.dob else "",
                s.gender,
                s.parent_name,
                s.contact_number,
            ])

        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 18

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="students.xlsx"'
        wb.save(response)
        return response