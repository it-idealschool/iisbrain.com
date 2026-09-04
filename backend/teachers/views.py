from rest_framework import viewsets, permissions, filters
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from .models import Teacher
from .serializers import TeacherSerializer
from sitesettings.models import RegistrationSettings


class TeacherViewSet(viewsets.ModelViewSet):
    serializer_class = TeacherSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'emp_no', 'qatar_id', 'email', 'position']

    def get_permissions(self):
        # Public self-registration: anyone can submit a new teacher record
        # (subject to the admin's on/off toggle, checked in create()).
        # Viewing/editing/deleting the list still requires login.
        if self.action == 'create':
            return [permissions.AllowAny()]
        return super().get_permissions()

    def create(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            settings_row = RegistrationSettings.load()
            if not settings_row.teacher_registration_open:
                raise PermissionDenied('Teacher registration is currently closed.')
        return super().create(request, *args, **kwargs)

    def get_queryset(self):
        qs = Teacher.objects.all().prefetch_related('grade_divisions', 'subject_periods__subject')

        params = self.request.query_params
        position = params.get('position')
        section = params.get('section')
        gender = params.get('gender')
        continue_service = params.get('continue_service')
        class_teacher = params.get('class_teacher')

        if position:
            qs = qs.filter(position=position)
        if section:
            qs = qs.filter(section=section)
        if gender:
            qs = qs.filter(gender=gender)
        if continue_service:
            qs = qs.filter(continue_service=continue_service)
        if class_teacher:
            qs = qs.filter(class_teacher=class_teacher)

        return qs

    @action(detail=False, methods=['get'], url_path='export')
    def export_excel(self, request):
        """
        Current search + filter query params-ന് അനുസരിച്ചുള്ള teacher list
        Excel (.xlsx) ആയി download ചെയ്യുന്നു.
        """
        queryset = self.filter_queryset(self.get_queryset())

        wb = Workbook()
        ws = wb.active
        ws.title = "Teachers"

        headers = [
            "Name", "Emp No", "Qatar ID", "Email", "Contact Number",
            "Position", "Section", "Gender", "Date of Joining",
            "Contract Expiry", "Total Periods", "Class Teacher",
            "Continue Service",
        ]
        ws.append(headers)

        for t in queryset:
            ws.append([
                t.name,
                t.emp_no,
                t.qatar_id,
                t.email,
                t.contact_number,
                t.position,
                t.section,
                t.gender,
                t.doj.isoformat() if t.doj else "",
                t.contract_expiry.isoformat() if t.contract_expiry else "",
                t.total_periods,
                t.class_teacher,
                t.continue_service,
            ])

        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 18

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="teachers.xlsx"'
        wb.save(response)
        return response